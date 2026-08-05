import io
import os
import time

os.environ["ENABLE_MOCK_AMAP"]="true"
os.environ["DATABASE_URL"]="sqlite://"
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.main import app


def test_full_mock_flow():
    with TestClient(app) as client:
        r=client.post("/api/stores/search",json={"name":"零食很忙武汉南湖店","city":"武汉市","district":"洪山区","address":""})
        assert r.status_code==200 and len(r.json()["data"]["candidates"])>=2
        c=r.json()["data"]["candidates"][0]
        store=client.post("/api/stores",json={"name":"零食很忙武汉南湖店","city":"武汉市","district":"洪山区","address":"","candidate":c}).json()["data"]
        assert client.post(f"/api/stores/{store['id']}/confirm",json={"candidate":c}).status_code==200
        analysis=client.post(f"/api/stores/{store['id']}/poi-search",json={"categories":["住宅小区","幼儿园","竞品门店"],"radii":[500,1000,2000],"max_radius":2000})
        assert analysis.status_code==200
        body=analysis.json()["data"]
        assert body["pois"] and len({p["id"] for p in body["pois"]})==len(body["pois"])

        profile = client.post(
            f"/api/stores/{store['id']}/business-district-analysis",
            json={"radii": [500, 1000, 2000]},
        )
        assert profile.status_code == 200
        profile_data = profile.json()["data"]
        assert profile_data["business_area"]["name"] == "南湖商圈"
        assert profile_data["business_district_type"]
        assert profile_data["level"]["level"] in {"S", "A", "B", "C", "D"}
        assert profile_data["analysis_version"]

        batch_profile = client.post(
            f"/api/analysis-jobs/{body['job_id']}/business-district-analysis",
            json={"radii": [500, 1000, 2000]},
        )
        assert batch_profile.status_code == 200
        compared = client.get(f"/api/analysis-jobs/{body['job_id']}/business-district-results")
        assert compared.status_code == 200 and len(compared.json()["data"]) == 1

        export=client.get(f"/api/analysis-jobs/{body['job_id']}/export")
        assert export.status_code==200
        wb=load_workbook(io.BytesIO(export.content))
        assert {"门店汇总","POI明细","待确认门店","失败记录","搜索配置","商圈汇总","商圈特征明细","潜在人群画像","商圈评分配置"}<=set(wb.sheetnames)

        business_export = client.get(f"/api/analysis-jobs/{body['job_id']}/business-district-export")
        business_wb = load_workbook(io.BytesIO(business_export.content))
        assert {"商圈汇总", "商圈特征明细", "潜在人群画像", "商圈评分配置"} == set(business_wb.sheetnames)


def test_import_preview_and_errors():
    with TestClient(app) as client:
        data="门店名称,城市,区县\n测试门店,武汉市,洪山区\n".encode("utf-8-sig")
        r=client.post("/api/import/preview",files={"file":("stores.csv",data,"text/csv")})
        assert r.status_code==200 and r.json()["data"]["mapping"]["name"]=="门店名称"
        r=client.post("/api/import/preview",files={"file":("bad.exe",b"x","application/octet-stream")})
        assert r.status_code==400 and r.json()["success"] is False

        address_only = "详细地址,城市\n武汉市洪山区文治街32号,武汉市\n".encode("utf-8-sig")
        r = client.post("/api/import/preview", files={"file": ("addresses.csv", address_only, "text/csv")})
        assert r.status_code == 200 and r.json()["data"]["mapping"]["address"] == "详细地址"


def test_batch_import_matching_confirmation_and_profile_flow():
    with TestClient(app) as client:
        rows = [
            {"门店名称":"零食很忙武汉南湖店","城市":"武汉市","区县":"洪山区","详细地址":""},
            {"门店名称":"零食很忙武汉光谷店","城市":"武汉市","区县":"洪山区","详细地址":""},
        ]
        created = client.post("/api/import/confirm", json={
            "filename":"batch.csv",
            "mapping":{"name":"门店名称","city":"城市","district":"区县","address":"详细地址"},
            "rows":rows,
            "config":{"radii":[500,1000],"categories":["住宅小区","幼儿园","小学"],"generate_profile":True},
        })
        assert created.status_code == 200
        job_id = created.json()["data"]["job_id"]
        job = client.get(f"/api/analysis-jobs/{job_id}").json()["data"]
        assert job["status"] == "等待开始匹配"
        assert job["config"]["radii"] == [500,1000]

        matched = client.post(f"/api/analysis-jobs/{job_id}/match-next", json={"batch_size":10})
        assert matched.status_code == 200
        match_data = matched.json()["data"]
        if match_data["pending"]:
            pending = client.get(f"/api/analysis-jobs/{job_id}/pending-matches").json()["data"]
            selections = [
                {"store_id":item["store"]["id"],"candidate_id":item["candidates"][0]["id"]}
                for item in pending if item["candidates"]
            ]
            confirmed = client.post(f"/api/analysis-jobs/{job_id}/confirm-matches", json={"selections":selections})
            assert confirmed.status_code == 200
        profile = client.post(f"/api/analysis-jobs/{job_id}/business-district-analysis", json={"radii":[500,1000]})
        assert profile.status_code == 200
        results = client.get(f"/api/analysis-jobs/{job_id}/business-district-results").json()["data"]
        assert len(results) == 2
        assert all(item["audience_profile"] for item in results)
        assert all(item["poi_summary"]["total"] > 0 for item in results)

        stores = client.get(f"/api/analysis-jobs/{job_id}/stores")
        assert stores.status_code == 200 and len(stores.json()["data"]) == 2
        first_store_id = stores.json()["data"][0]["store"]["id"]
        detail = client.get(f"/api/analysis-jobs/{job_id}/stores/{first_store_id}")
        assert detail.status_code == 200
        detail_data = detail.json()["data"]
        assert detail_data["poi_summary"]["total"] == len(detail_data["pois"])
        assert set(detail_data["poi_summary"]["by_category"]) <= {"住宅小区", "幼儿园", "小学"}
        assert detail_data["analysis"]["audience_profile"]

        store_export = client.get(f"/api/analysis-jobs/{job_id}/stores/{first_store_id}/export")
        assert store_export.status_code == 200
        store_wb = load_workbook(io.BytesIO(store_export.content))
        assert {"门店概览", "POI明细", "商圈汇总", "商圈特征明细", "潜在人群画像", "商圈评分配置"} == set(store_wb.sheetnames)

        rerun = client.post(f"/api/analysis-jobs/{job_id}/business-district-analysis", json={"radii":[500,1000]})
        assert rerun.status_code == 200
        rerun_results = client.get(f"/api/analysis-jobs/{job_id}/business-district-results").json()["data"]
        assert len(rerun_results) == 2


def test_address_geocode_and_config_roundtrip():
    with TestClient(app) as client:
        matched = client.post(
            "/api/stores/search",
            json={"name": "零食很忙武汉南湖店", "city": "武汉市", "district": "洪山区", "address": "文治街32号"},
        )
        assert matched.status_code == 200
        assert matched.json()["data"]["candidates"][0]["address_geocode_distance_m"] == 0

        response = client.post(
            "/api/geocode",
            json={"address": "文治街32号", "province": "湖北省", "city": "武汉市", "district": "洪山区"},
        )
        assert response.status_code == 200
        candidate = response.json()["data"]["candidates"][0]
        assert candidate["level"] == "门牌号"
        assert candidate["requires_confirmation"] is False

        created = client.post(
            "/api/stores/from-geocode",
            json={"name": "地址定位测试店", "address": "文治街32号", "city": "武汉市", "candidate": candidate},
        )
        assert created.status_code == 200
        store = created.json()["data"]
        assert store["raw_address"] == "文治街32号"
        assert store["location_source"] == "高德地址解析"

        config = client.get("/api/business-district-config")
        assert config.status_code == 200
        assert config.json()["data"]["analysis_version"]
        saved = client.put("/api/business-district-config", json={"version": "test-v1", "note": "测试配置"})
        assert saved.status_code == 200 and saved.json()["data"]["version"] == "test-v1"


def test_ipv6_local_frontend_origin_is_allowed():
    with TestClient(app) as client:
        r = client.options(
            "/api/health",
            headers={
                "Origin": "http://[::1]:3000",
                "Access-Control-Request-Method": "GET",
            },
        )
        assert r.status_code == 200
        assert r.headers["access-control-allow-origin"] == "http://[::1]:3000"


def test_background_batch_matching_analysis_and_controls():
    with TestClient(app) as client:
        rows = [
            {"门店名称": f"零食很忙武汉测试店{i}", "城市": "武汉市", "区县": "洪山区", "详细地址": ""}
            for i in range(1, 11)
        ]
        created = client.post("/api/import/confirm", json={
            "filename": "ten-stores.xlsx",
            "mapping": {"name": "门店名称", "city": "城市", "district": "区县", "address": "详细地址"},
            "rows": rows,
            "config": {"radii": [500, 1000], "categories": ["住宅小区", "幼儿园", "小学"], "generate_profile": True},
        })
        job_id = created.json()["data"]["job_id"]

        paused = client.post(f"/api/analysis-jobs/{job_id}/pause")
        assert paused.status_code == 200 and paused.json()["data"]["status"] == "已暂停"
        resumed = client.post(f"/api/analysis-jobs/{job_id}/resume")
        assert resumed.status_code == 200
        for _ in range(100):
            job = client.get(f"/api/analysis-jobs/{job_id}").json()["data"]
            if job["status"] in {"匹配完成", "匹配部分失败"}:
                break
            time.sleep(0.02)
        assert job["matched_stores"] == 10 and job["pending_stores"] == 0
        assert job["progress_percent"] == 100

        started = client.post(f"/api/analysis-jobs/{job_id}/start-analysis")
        assert started.status_code == 200
        duplicate_start = client.post(f"/api/analysis-jobs/{job_id}/start-analysis")
        assert duplicate_start.status_code == 200
        for _ in range(150):
            job = client.get(f"/api/analysis-jobs/{job_id}").json()["data"]
            if job["status"] in {"已完成", "部分完成"}:
                break
            time.sleep(0.02)
        assert job["success_stores"] == 10
        assert job["stage"] == "analysis" and job["progress_percent"] == 100

        ended = client.post(f"/api/analysis-jobs/{job_id}/end")
        assert ended.status_code == 200 and ended.json()["data"]["status"] == "已结束"
