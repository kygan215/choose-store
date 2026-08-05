from __future__ import annotations

import asyncio
import io
import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, model_validator
from sqlalchemy import select

from .amap import AmapClient, AmapServiceError, CATEGORIES
from .business_district import DISCLAIMER, FEATURE_PACK, analyze_business_district, public_config
from .core import confidence_status, deduplicate_pois, distance_bucket, evaluate_candidate, haversine_m, map_headers, safe_excel
from .database import AnalysisJob, AnalysisJobStore, AuditLog, BusinessDistrictAnalysis, BusinessDistrictConfig, PoiCategory, PoiResult, SearchRequestLog, SessionLocal, Store, StoreMatchCandidate, init_db

app = FastAPI(title="门店周边 POI 搜索与分析平台", version="1.2.0")
JOB_RUNNERS: dict[int, asyncio.Task[Any]] = {}
MAX_CONCURRENT_STORES = max(1, min(5, int(os.getenv("MAX_CONCURRENT_STORES", os.getenv("AMAP_MAX_CONCURRENCY", "3")))))
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://[::1]:3000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup() -> None:
    init_db()
    with SessionLocal() as db:
        if not db.scalar(select(PoiCategory.id).limit(1)):
            for i, (name, (codes, keywords)) in enumerate(CATEGORIES.items()):
                db.add(PoiCategory(name=name, display_name=name, search_mode="keyword" if keywords else "typecode", typecodes=codes or None, keywords=keywords.split("|") if keywords else [], sort_order=i))
            db.commit()


@app.middleware("http")
async def request_context(request: Request, call_next):
    request.state.request_id = str(uuid.uuid4())
    response = await call_next(request)
    response.headers["X-Request-ID"] = request.state.request_id
    return response


def ok(data: Any = None, message: str = "操作成功", request_id: str | None = None):
    return {"success": True, "data": data, "message": message, "request_id": request_id or str(uuid.uuid4())}


class StoreQuery(BaseModel):
    name: str = Field(min_length=2, max_length=200)
    province: str = Field("", max_length=80)
    city: str = Field("", max_length=80)
    district: str = Field("", max_length=80)
    adcode: str = Field("", max_length=20)
    address: str = Field("", max_length=300)


class StoreCreate(StoreQuery):
    candidate: dict[str, Any] | None = None


class ConfirmBody(BaseModel):
    candidate: dict[str, Any]
    confirmation_method: str = "用户手动选择候选"
    confirmed_by: str = "本地用户"


class LocationBody(BaseModel):
    longitude: float = Field(ge=73, le=136)
    latitude: float = Field(ge=3, le=54)


class SearchBody(BaseModel):
    categories: list[str] = Field(min_length=1)
    radii: list[int] = Field(min_length=1)
    max_radius: int = Field(ge=1, le=50000)

    @model_validator(mode="after")
    def validate_radii(self):
        self.radii = sorted(set(self.radii))
        if len(self.radii) > 5:
            raise ValueError("最多同时选择 5 个搜索半径")
        if any(radius <= 0 or radius > 50000 for radius in self.radii):
            raise ValueError("搜索半径必须大于 0 且不超过 50 公里")
        if self.max_radius != max(self.radii):
            self.max_radius = max(self.radii)
        return self


class AddressQuery(BaseModel):
    province: str = Field("", max_length=80)
    city: str = Field("", max_length=80)
    district: str = Field("", max_length=80)
    address: str = Field(min_length=2, max_length=300)
    name: str = Field("", max_length=200)


class AddressConfirmBody(AddressQuery):
    candidate: dict[str, Any]


class BusinessAnalysisBody(BaseModel):
    radii: list[int] = Field(default_factory=lambda: [500])

    @model_validator(mode="after")
    def validate_business_radii(self):
        self.radii = sorted(set(self.radii))
        if not self.radii or len(self.radii) > 5 or any(x <= 0 or x > 50000 for x in self.radii):
            raise ValueError("商圈半径需为 1 至 5 个、且不超过 50 公里的正整数")
        return self


@app.get("/api/health")
def health():
    client = AmapClient()
    return ok({"mock": client.mock, "web_key": bool(client.key), "js_key": bool(os.getenv("NEXT_PUBLIC_AMAP_JS_KEY")), "coordinate_system": "GCJ-02"})


@app.post("/api/stores/search")
async def search_stores(body: StoreQuery):
    client = AmapClient()
    geocode_reference = None
    try:
        raw = await client.search_store(body.model_dump())
        if body.address:
            geocodes = await client.geocode(body.model_dump())
            geocode_reference = geocodes[0] if geocodes else None
    except AmapServiceError as exc:
        raise HTTPException(503, detail={"error_code": exc.code, "message": exc.message})
    query = body.model_dump()
    initial = [(item, evaluate_candidate(query, item)) for item in raw[:10]]
    high_matches = [result for _, result in initial if result["score"] >= 80 and not result["conflicts"]]
    unique_high_match = len(high_matches) == 1
    scored = []
    for item, _ in initial:
        result = evaluate_candidate(query, item, unique_high_match=unique_high_match)
        if geocode_reference:
            reference_location = geocode_reference.get("location") or []
            candidate_location = item.get("location") or []
            if len(reference_location) >= 2 and len(candidate_location) >= 2:
                cross_distance = haversine_m(tuple(reference_location[:2]), tuple(candidate_location[:2]))
                item["address_geocode_distance_m"] = cross_distance
                if cross_distance <= 800:
                    result["breakdown"].append({"label": f"地址解析坐标与候选相距 {cross_distance} 米，位置一致", "points": 0, "kind": "match"})
                elif cross_distance > 2000:
                    message = f"详细地址解析位置与候选相距 {cross_distance} 米"
                    result["conflicts"].append(message)
                    result["breakdown"].append({"label": message, "points": -10, "kind": "conflict"})
                    result["score"] = max(0, result["score"] - 10)
                    result["auto_confirm"] = False
                else:
                    result["warnings"].append(f"详细地址解析位置与候选相距 {cross_distance} 米，请核对地图")
        scored.append({**item, **result})
    scored.sort(key=lambda x: x["score"], reverse=True)
    for i, item in enumerate(scored):
        margin = item["score"] - (scored[i+1]["score"] if i + 1 < len(scored) else 0)
        item["status"] = "高置信度" if item["auto_confirm"] and margin >= 15 else confidence_status(item["score"], margin, bool(body.city))
    return ok({"candidates": scored, "mode": "mock" if client.mock else "amap"})


@app.post("/api/geocode")
async def geocode_address(body: AddressQuery):
    client = AmapClient()
    try:
        candidates = await client.geocode(body.model_dump())
    except AmapServiceError as exc:
        raise HTTPException(503, detail={"error_code": exc.code, "message": exc.message})
    low_precision = {"省", "市", "区县", "道路", "未知"}
    return ok({
        "candidates": [
            {**item, "requires_confirmation": item.get("level") in low_precision}
            for item in candidates
        ],
        "mode": "mock" if client.mock else "amap",
    })


@app.post("/api/stores/from-geocode")
def create_store_from_geocode(body: AddressConfirmBody):
    c = body.candidate
    location = c.get("location") or [None, None]
    if len(location) < 2 or location[0] is None or location[1] is None:
        raise HTTPException(400, "地址候选缺少有效坐标")
    with SessionLocal() as db:
        store = Store(
            input_name=body.name or body.address,
            standard_name=body.name or c.get("formatted_address"),
            longitude=location[0], latitude=location[1],
            province=c.get("province") or body.province, city=c.get("city") or body.city,
            district=c.get("district") or body.district, adcode=c.get("adcode"),
            address=c.get("formatted_address") or body.address, raw_address=body.address,
            standardized_address=c.get("formatted_address"), geocode_level=c.get("level"),
            location_source="高德地址解析", source="mock" if c.get("source") == "mock" else "amap",
            match_status="待确认",
        )
        db.add(store); db.commit()
        return ok(serialize_store(store))


@app.post("/api/stores")
def create_store(body: StoreCreate):
    with SessionLocal() as db:
        c = body.candidate or {}
        store = Store(input_name=body.name, standard_name=c.get("name"), city=body.city or c.get("city"), district=body.district or c.get("district"), address=body.address or c.get("address"), raw_address=body.address or None, standardized_address=c.get("formatted_address") or c.get("address"), geocode_level=c.get("level"), location_source=c.get("location_source") or "高德门店POI", source="mock" if str(c.get("id","")).startswith("MOCK") else "amap", match_score=c.get("score"), match_status="待确认")
        db.add(store); db.commit()
        return ok({"id": store.id})


@app.get("/api/stores")
def list_stores():
    with SessionLocal() as db:
        stores = db.scalars(select(Store).order_by(Store.created_at.desc())).all()
        return ok([serialize_store(s) for s in stores])


@app.get("/api/stores/{store_id}")
def get_store(store_id: int):
    with SessionLocal() as db:
        store = db.get(Store, store_id)
        if not store: raise HTTPException(404, "门店不存在")
        return ok(serialize_store(store))


@app.post("/api/stores/{store_id}/confirm")
def confirm_store(store_id: int, body: ConfirmBody):
    with SessionLocal() as db:
        store = db.get(Store, store_id)
        if not store: raise HTTPException(404, "门店不存在")
        c = body.candidate; loc = c.get("location") or [None, None]
        store.standard_name, store.amap_poi_id = c.get("name"), c.get("id")
        store.longitude, store.latitude = loc[0], loc[1]
        store.province, store.city, store.district = c.get("province"), c.get("city"), c.get("district")
        store.adcode, store.address = c.get("adcode"), c.get("address")
        store.poi_type, store.poi_typecode = c.get("type"), c.get("typecode")
        store.match_score, store.match_status = c.get("score"), "已确认"
        store.location_source = "用户候选确认"
        store.confirmation_method, store.confirmed_by, store.confirmed_at = body.confirmation_method, body.confirmed_by, datetime.utcnow()
        store.last_verified_at = datetime.utcnow()
        db.add(AuditLog(action="确认门店", entity_type="Store", entity_id=store.id, actor=body.confirmed_by, details={"method": body.confirmation_method}))
        db.commit()
        return ok(serialize_store(store), "门店位置已确认")


@app.post("/api/stores/{store_id}/correct-location")
def correct_location(store_id: int, body: LocationBody):
    with SessionLocal() as db:
        store = db.get(Store, store_id)
        if not store: raise HTTPException(404, "门店不存在")
        store.longitude, store.latitude = body.longitude, body.latitude
        store.confirmation_method, store.confirmed_at = "用户地图拖动修正", datetime.utcnow()
        store.location_source = "用户地图修正"
        db.add(AuditLog(action="修正坐标", entity_type="Store", entity_id=store.id, details=body.model_dump()))
        db.commit()
        return ok(serialize_store(store))


@app.get("/api/poi-categories")
def list_categories():
    with SessionLocal() as db:
        items = db.scalars(select(PoiCategory).where(PoiCategory.enabled == True).order_by(PoiCategory.sort_order)).all()
        return ok([{"id":x.id,"name":x.name,"display_name":x.display_name,"search_mode":x.search_mode,"typecodes":x.typecodes,"keywords":x.keywords,"color":x.color} for x in items])


@app.post("/api/poi-categories")
def create_category(body: dict[str, Any]):
    with SessionLocal() as db:
        cat = PoiCategory(**{k:v for k,v in body.items() if hasattr(PoiCategory,k)})
        db.add(cat); db.commit()
        return ok({"id":cat.id})


@app.put("/api/poi-categories/{category_id}")
def update_category(category_id: int, body: dict[str, Any]):
    with SessionLocal() as db:
        cat = db.get(PoiCategory, category_id)
        if not cat: raise HTTPException(404, "分类不存在")
        for k,v in body.items():
            if k not in {"id","created_at"} and hasattr(cat,k): setattr(cat,k,v)
        db.commit(); return ok({"id":cat.id})


@app.post("/api/stores/{store_id}/poi-search")
async def poi_search(store_id: int, body: SearchBody):
    with SessionLocal() as db:
        store = db.get(Store, store_id)
        if not store or store.longitude is None: raise HTTPException(400, "请先确认门店位置")
        job = AnalysisJob(status="正在查询POI", total_stores=1, processed_stores=0, matched_stores=1, config=body.model_dump())
        db.add(job); db.flush()
        db.add(AnalysisJobStore(analysis_job_id=job.id, store_id=store.id, status="正在查询POI"))
        db.commit(); job_id = job.id
    client, all_items, truncated = AmapClient(), [], False
    try:
        for category in body.categories:
            items, cut = await client.around((store.longitude, store.latitude), category, body.max_radius)
            truncated = truncated or cut
            for item in items:
                if item["distance"] <= body.max_radius:
                    all_items.append({**item, "category": category})
    except AmapServiceError as exc:
        with SessionLocal() as db:
            job = db.get(AnalysisJob, job_id); job.status = "失败"; job.failed_stores = 1; db.commit()
        raise HTTPException(503, detail={"error_code": exc.code, "message": exc.message})
    all_items = deduplicate_pois(all_items)
    with SessionLocal() as db:
        for item in all_items:
            db.add(PoiResult(analysis_job_id=job_id, store_id=store_id, amap_poi_id=item["id"], parent_poi_id=item.get("parent"), name=item["name"], poi_category=item["category"], poi_type=item.get("type"), poi_typecode=item.get("typecode"), longitude=item["location"][0], latitude=item["location"][1], straight_distance_m=item["distance"], province=item.get("province"), city=item.get("city"), district=item.get("district"), adcode=item.get("adcode"), address=item.get("address"), business_area=item.get("business_area"), distance_bucket=distance_bucket(item["distance"], body.radii), search_keyword=item["category"] if item["category"] == "竞品门店" else None, search_typecodes=CATEGORIES.get(item["category"],("", ""))[0], search_radius=body.max_radius))
        job = db.get(AnalysisJob, job_id); job.status = "已完成"; job.processed_stores = job.success_stores = 1; job.truncated = truncated
        link = db.scalar(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id==job_id)); link.status="已完成"
        db.add(SearchRequestLog(request_id=str(uuid.uuid4()), endpoint="/v5/place/around", params={"categories":body.categories,"radii":body.radii,"radius":body.max_radius}, status="成功", truncated=truncated))
        db.commit()
    return ok({"job_id":job_id,"pois":[{"id":x["id"],"name":x["name"],"category":x["category"],"type":x.get("type"),"address":x.get("address"),"distance":x["distance"],"location":x["location"],"distance_bucket":distance_bucket(x["distance"],body.radii)} for x in sorted(all_items,key=lambda x:x["distance"])],"truncated":truncated})


@app.get("/api/stores/{store_id}/poi-results")
def poi_results(store_id: int, limit: int = 200, offset: int = 0):
    with SessionLocal() as db:
        items = db.scalars(select(PoiResult).where(PoiResult.store_id==store_id,PoiResult.excluded==False).order_by(PoiResult.straight_distance_m).offset(offset).limit(min(limit,500))).all()
        return ok([serialize_poi(x) for x in items])


@app.get("/api/stores/{store_id}/summary")
def store_summary(store_id: int):
    with SessionLocal() as db:
        items = db.scalars(select(PoiResult).where(PoiResult.store_id==store_id,PoiResult.excluded==False)).all()
        by_category, by_distance = {}, {}
        for x in items:
            by_category[x.poi_category] = by_category.get(x.poi_category,0)+1
            by_distance[x.distance_bucket] = by_distance.get(x.distance_bucket,0)+1
        return ok({"total":len(items),"by_category":by_category,"by_distance":by_distance,"disclaimer":"本分析仅反映周边设施和兴趣点分布，不等同于人口、客流、消费能力或销售预测。"})


def _serialize_business_analysis(item: BusinessDistrictAnalysis) -> dict[str, Any]:
    return {
        "id": item.id,
        "store_id": item.store_id,
        "analysis_job_id": item.analysis_job_id,
        "analysis_version": item.analysis_version,
        "poi_config_version": item.poi_config_version,
        "weight_version": item.weight_version,
        "amap_query_time": item.amap_query_time,
        "radius_config": item.radius_config,
        "location_source": item.location_source,
        "business_area": {
            "name": item.business_area_name,
            "source": item.business_area_source,
            "confidence": item.business_area_confidence,
        },
        "business_district_type": {
            "type": item.business_district_type,
            "scores": item.type_scores,
            "confidence": item.type_confidence,
        },
        "level": {"level": item.level, "score": item.level_score, "mode": item.level_mode},
        "fit": {"score": item.fit_score, "level": item.fit_level},
        "competition": {"score": item.competition_score, "level": item.competition_level},
        "audience_profile": (item.feature_vector or {}).get("audience_profile"),
        "confidence_level": item.confidence_level,
        "feature_vector": item.feature_vector,
        "evidence": item.evidence,
        "strengths": item.strengths,
        "weaknesses": item.weaknesses,
        "truncation_flags": item.truncation_flags,
        "warning_messages": item.warning_messages,
        "disclaimer": DISCLAIMER,
        "created_at": item.created_at.isoformat(),
        "updated_at": item.updated_at.isoformat(),
    }


async def _run_business_analysis(store_id: int, radii: list[int], analysis_job_id: int | None = None) -> BusinessDistrictAnalysis:
    max_radius = max(radii)
    with SessionLocal() as db:
        store = db.get(Store, store_id)
        if not store or store.longitude is None or store.latitude is None:
            raise HTTPException(400, "请先确认门店位置")
        cached_rows = db.scalars(select(PoiResult).where(PoiResult.store_id == store_id, PoiResult.excluded == False, PoiResult.straight_distance_m <= max_radius)).all()
        cached_categories = {row.poi_category for row in cached_rows if row.poi_category in FEATURE_PACK}
        if analysis_job_id is None:
            job = AnalysisJob(status="正在生成商圈画像", total_stores=1, matched_stores=1, config={"business_district": True, "radii": radii})
            db.add(job); db.flush()
            db.add(AnalysisJobStore(analysis_job_id=job.id, store_id=store_id, status="正在生成商圈画像"))
            analysis_job_id = job.id
            db.commit()

    failures: list[str] = []
    truncated = False
    client = AmapClient()
    new_items: list[dict[str, Any]] = []
    for category in FEATURE_PACK:
        if category in cached_categories:
            continue
        try:
            items, cut = await client.around((store.longitude, store.latitude), category, max_radius)
            truncated = truncated or cut
            new_items.extend({**item, "category": category} for item in items if item["distance"] <= max_radius)
        except AmapServiceError as exc:
            failures.append(f"{category}查询失败：{exc.message}")

    with SessionLocal() as db:
        for item in deduplicate_pois(new_items):
            exists = db.scalar(select(PoiResult.id).where(PoiResult.store_id == store_id, PoiResult.amap_poi_id == item["id"], PoiResult.poi_category == item["category"]).limit(1))
            if exists:
                continue
            db.add(PoiResult(
                analysis_job_id=analysis_job_id, store_id=store_id, amap_poi_id=item["id"], parent_poi_id=item.get("parent"),
                name=item["name"], poi_category=item["category"], poi_type=item.get("type"), poi_typecode=item.get("typecode"),
                longitude=item["location"][0], latitude=item["location"][1], straight_distance_m=item["distance"],
                province=item.get("province"), city=item.get("city"), district=item.get("district"), adcode=item.get("adcode"),
                address=item.get("address"), business_area=item.get("business_area"), distance_bucket=distance_bucket(item["distance"], radii),
                search_keyword=FEATURE_PACK[item["category"]]["keywords"] or None, search_typecodes=FEATURE_PACK[item["category"]]["types"] or None,
                search_radius=max_radius,
            ))
        db.commit()
        rows = db.scalars(select(PoiResult).where(PoiResult.store_id == store_id, PoiResult.excluded == False, PoiResult.straight_distance_m <= max_radius, PoiResult.poi_category.in_(list(FEATURE_PACK)))).all()
        data_items = [{"category": row.poi_category, "distance": row.straight_distance_m, "business_area": row.business_area, "name": row.name} for row in rows]
        result = analyze_business_district(
            data_items, radii=radii, location_confirmed=bool(store.confirmed_at), truncated=truncated, failures=failures,
        )
        area, type_result, level, fit, competition = result["business_area"], result["business_district_type"], result["level"], result["fit"], result["competition"]
        record = BusinessDistrictAnalysis(
            store_id=store_id, analysis_job_id=analysis_job_id, analysis_version=result["analysis_version"],
            poi_config_version=result["poi_config_version"], weight_version=result["weight_version"], amap_query_time=result["amap_query_time"],
            radius_config=result["radius_config"], location_source=store.location_source or store.confirmation_method,
            business_area_name=area["name"], business_area_source=area["source"], business_area_confidence=area["confidence"],
            business_district_type=type_result["type"], type_scores=type_result["scores"], type_confidence=type_result["confidence"],
            level=level["level"], level_score=level["score"], level_mode=level["mode"], fit_score=fit["score"], fit_level=fit["level"],
            competition_score=competition["score"], competition_level=competition["level"], confidence_level=result["confidence_level"],
            feature_vector=result["feature_vector"], evidence=result["evidence"], strengths=result["strengths"], weaknesses=result["weaknesses"],
            truncation_flags=result["truncation_flags"], warning_messages=result["warning_messages"],
        )
        for old in db.scalars(select(BusinessDistrictAnalysis).where(
            BusinessDistrictAnalysis.analysis_job_id == analysis_job_id,
            BusinessDistrictAnalysis.store_id == store_id,
        )).all():
            db.delete(old)
        db.flush()
        db.add(record)
        job = db.get(AnalysisJob, analysis_job_id)
        if job:
            job.status = "已完成" if not failures else "部分完成"
            job.processed_stores = job.success_stores = 1
            job.failed_stores = 0
            job.truncated = truncated
        link = db.scalar(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id == analysis_job_id, AnalysisJobStore.store_id == store_id))
        if link:
            link.status = "分析完成" if not failures else "分析失败"
            link.error_message = "；".join(failures) or None
        db.commit(); db.refresh(record)
        return record


@app.post("/api/stores/{store_id}/business-district-analysis")
async def create_business_analysis(store_id: int, body: BusinessAnalysisBody):
    record = await _run_business_analysis(store_id, body.radii)
    return ok(_serialize_business_analysis(record), "商圈画像已生成")


@app.get("/api/stores/{store_id}/business-district-analysis")
def get_business_analysis(store_id: int):
    with SessionLocal() as db:
        item = db.scalar(select(BusinessDistrictAnalysis).where(BusinessDistrictAnalysis.store_id == store_id).order_by(BusinessDistrictAnalysis.created_at.desc()).limit(1))
        if not item:
            raise HTTPException(404, "该门店尚未生成商圈画像")
        return ok(_serialize_business_analysis(item))


@app.get("/api/business-district-config")
def get_business_config():
    with SessionLocal() as db:
        custom = db.scalar(select(BusinessDistrictConfig).where(BusinessDistrictConfig.active == True).order_by(BusinessDistrictConfig.updated_at.desc()).limit(1))
        config = public_config()
        if custom:
            config["custom"] = custom.config
        return ok(config)


@app.put("/api/business-district-config")
def update_business_config(body: dict[str, Any]):
    version = str(body.get("version") or f"custom-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}")
    with SessionLocal() as db:
        for item in db.scalars(select(BusinessDistrictConfig).where(BusinessDistrictConfig.active == True)).all():
            item.active = False
        record = db.scalar(select(BusinessDistrictConfig).where(BusinessDistrictConfig.version == version))
        if record:
            record.config = body
            record.active = True
        else:
            record = BusinessDistrictConfig(version=version, config=body, active=True)
            db.add(record)
        db.commit()
        return ok({"version": version})


MAX_UPLOAD = 15 * 1024 * 1024


def _validate_import_rows(rows: list[dict[str, Any]], mapping: dict[str, str]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int, list[dict[str, Any]]]:
    valid, preview, validated, duplicate_count = [], [], [], 0
    seen: set[tuple[str, ...]] = set()
    for index, row in enumerate(rows, start=2):
        name = str(row.get(mapping.get("name", ""), "")).strip()
        address = str(row.get(mapping.get("address", ""), "")).strip()
        city = str(row.get(mapping.get("city", ""), "")).strip()
        code = str(row.get(mapping.get("code", ""), "")).strip()
        issues: list[str] = []
        if not name and not address:
            issues.append("门店名称和详细地址不能同时为空")
        key = ("code", code.lower()) if code else ("store", name.lower(), city.lower(), address.lower())
        if key in seen:
            issues.append("与文件内其他门店重复")
            duplicate_count += 1
        else:
            seen.add(key)
        enriched = {**row, "_row_number": index, "_valid": not issues, "_issues": issues}
        validated.append(enriched)
        if len(preview) < 20:
            preview.append(enriched)
        if not issues:
            valid.append(row)
    return valid, preview, duplicate_count, validated


@app.post("/api/import/preview")
async def import_preview(file: UploadFile = File(...)):
    filename = Path(file.filename or "").name
    ext = Path(filename).suffix.lower()
    if ext not in {".xlsx",".xls",".csv"}: raise HTTPException(400, "仅支持 .xlsx、.xls、.csv 文件")
    content = await file.read(MAX_UPLOAD + 1)
    if len(content) > MAX_UPLOAD: raise HTTPException(413, "文件过大，最大允许 15 MB")
    try:
        if ext == ".csv":
            for encoding in ("utf-8-sig","gb18030"):
                try: df = pd.read_csv(io.BytesIO(content),encoding=encoding,dtype=str); break
                except UnicodeDecodeError: continue
        else: df = pd.read_excel(io.BytesIO(content),dtype=str)
    except Exception as exc:
        raise HTTPException(400, "文件无法解析，请检查文件是否损坏或加密") from exc
    df = df.dropna(how="all")
    if len(df) > int(os.getenv("MAX_IMPORT_ROWS","5000")): raise HTTPException(400, "导入行数超过 5,000 行限制")
    headers = [str(x).strip() for x in df.columns]
    mapping = map_headers(headers)
    if "name" not in mapping and "address" not in mapping:
        raise HTTPException(400, "门店名称和详细地址至少需要识别到一列")
    clean_df = df.fillna("").astype(str)
    all_rows = clean_df.to_dict("records")
    valid_rows, rows, duplicates, selectable_rows = _validate_import_rows(all_rows, mapping)
    warnings = []
    if duplicates:
        warnings.append(f"发现 {duplicates} 行重复门店，创建任务时会自动跳过")
    invalid = len(all_rows) - len(valid_rows) - duplicates
    if invalid > 0:
        warnings.append(f"发现 {invalid} 行缺少门店名称和详细地址")
    return ok({
        "filename":filename,"file_size":len(content),"headers":headers,"mapping":mapping,"rows":rows,"all_rows":all_rows,"selectable_rows":selectable_rows,
        "total_rows":len(df),"valid_rows":len(valid_rows),"invalid_rows":len(all_rows)-len(valid_rows),
        "duplicate_rows":duplicates,"warnings":warnings,
    })


@app.get("/api/import/template")
def import_template():
    wb = Workbook(); ws = wb.active; ws.title = "门店导入模板"
    headers = ["门店名称","省份","城市","区县","详细地址","门店编号","品牌","备注"]
    ws.append(headers); ws.append(["零食很忙武汉南湖店","湖北省","武汉市","洪山区","文治街32号","WH001","零食很忙","演示数据"])
    style_sheet(ws)
    buf=io.BytesIO();wb.save(buf);buf.seek(0)
    return StreamingResponse(buf,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":"attachment; filename=store_import_template.xlsx"})


@app.post("/api/import/confirm")
def import_confirm(body: dict[str, Any]):
    rows = body.get("rows") or []
    mapping = body.get("mapping") or {}
    if "name" not in mapping and "address" not in mapping:
        raise HTTPException(400, "请至少映射门店名称或详细地址")
    valid_rows, _, duplicates, _ = _validate_import_rows(rows, mapping)
    if not valid_rows:
        raise HTTPException(400, "没有可导入的有效门店")
    config = body.get("config") or {}
    radii = sorted(set(int(x) for x in (config.get("radii") or [500])))
    categories = [x for x in (config.get("categories") or ["住宅小区","幼儿园","小学"]) if x in CATEGORIES]
    if not radii or len(radii) > 5 or any(x <= 0 or x > 50000 for x in radii):
        raise HTTPException(400, "批量搜索半径需为 1 至 5 个、且不超过 50 公里的正整数")
    if not categories:
        raise HTTPException(400, "请至少选择一个 POI 分类")
    with SessionLocal() as db:
        job=AnalysisJob(filename=body.get("filename"),status="等待开始匹配",total_stores=len(valid_rows),config={"mapping":mapping,"radii":radii,"categories":categories,"generate_profile":bool(config.get("generate_profile",True)),"active_stage":"match","control":"idle","stage_total":len(valid_rows),"stage_processed":0,"current_store":""});db.add(job);db.flush()
        for row in valid_rows:
            name = str(row.get(mapping.get("name", ""), "")).strip()
            address = str(row.get(mapping.get("address", ""), "")).strip()
            store=Store(input_name=name or address, user_code=str(row.get(mapping.get("code",""),"")) or None, province=str(row.get(mapping.get("province",""),"")), city=str(row.get(mapping.get("city",""),"")),district=str(row.get(mapping.get("district",""),"")),address=address,raw_address=address or None,location_source="待定位");db.add(store);db.flush()
            db.add(AnalysisJobStore(analysis_job_id=job.id,store_id=store.id,status="等待匹配"))
        db.commit()
        return ok({"job_id":job.id,"status":job.status,"accepted":len(valid_rows),"rejected":len(rows)-len(valid_rows),"duplicates":duplicates})


@app.get("/api/analysis-jobs")
def list_jobs():
    with SessionLocal() as db:
        jobs=db.scalars(select(AnalysisJob).order_by(AnalysisJob.created_at.desc())).all()
        return ok([serialize_job(x) for x in jobs])


@app.get("/api/analysis-jobs/{job_id}")
def get_job(job_id:int):
    with SessionLocal() as db:
        job=db.get(AnalysisJob,job_id)
        if not job: raise HTTPException(404,"任务不存在")
        return ok(serialize_job(job))


@app.post("/api/analysis-jobs/{job_id}/resume")
async def resume(job_id:int):
    with SessionLocal() as db:
        job=db.get(AnalysisJob,job_id)
        if not job: raise HTTPException(404,"任务不存在")
        config=dict(job.config or {})
        stage=str(config.get("active_stage") or "match")
        config["control"]="run";job.config=config
        job.status="正在匹配门店" if stage=="match" else "正在完整分析"
        db.commit();payload=serialize_job(job)
    _schedule_job_runner(job_id,stage)
    return ok(payload,"任务已继续")


@app.post("/api/analysis-jobs/{job_id}/pause")
def pause_job(job_id:int):
    with SessionLocal() as db:
        job=db.get(AnalysisJob,job_id)
        if not job: raise HTTPException(404,"任务不存在")
        config=dict(job.config or {});config["control"]="pause";job.config=config;job.status="已暂停"
        db.commit();return ok(serialize_job(job),"任务将在当前门店处理完成后暂停")


@app.post("/api/analysis-jobs/{job_id}/end")
def end_job(job_id:int):
    with SessionLocal() as db:
        job=db.get(AnalysisJob,job_id)
        if not job: raise HTTPException(404,"任务不存在")
        config=dict(job.config or {});config["control"]="stop";config["current_store"]="";job.config=config;job.status="已结束"
        db.commit();return ok(serialize_job(job),"任务已结束，已完成结果将保留")


@app.post("/api/analysis-jobs/{job_id}/start-matching")
async def start_matching_all(job_id:int):
    with SessionLocal() as db:
        job=db.get(AnalysisJob,job_id)
        if not job: raise HTTPException(404,"任务不存在")
        if job_id in JOB_RUNNERS and not JOB_RUNNERS[job_id].done():
            return ok(serialize_job(job),"任务已在运行")
        config=dict(job.config or {});config.update({"active_stage":"match","control":"run","stage_total":job.total_stores,"stage_processed":job.processed_stores,"current_store":"准备开始匹配"});job.config=config;job.status="正在匹配门店"
        db.commit();payload=serialize_job(job)
    _schedule_job_runner(job_id,"match")
    return ok(payload,"已开始连续匹配全部门店")


@app.post("/api/analysis-jobs/{job_id}/start-analysis")
async def start_analysis_all(job_id:int):
    with SessionLocal() as db:
        job=db.get(AnalysisJob,job_id)
        if not job: raise HTTPException(404,"任务不存在")
        if job_id in JOB_RUNNERS and not JOB_RUNNERS[job_id].done():
            return ok(serialize_job(job),"任务已在运行")
        links=db.scalars(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id==job_id)).all()
        analyzable=sum(link.status in {"已确认","已完成","分析完成","分析失败"} for link in links)
        if not analyzable: raise HTTPException(400,"当前任务没有已匹配门店")
        for link in links:
            store=db.get(Store,link.store_id)
            if store and store.longitude is not None and link.status in {"已确认","已完成","分析完成","分析失败"}:
                link.status="已确认";link.error_code=None;link.error_message=None
        config=dict(job.config or {});config.update({"active_stage":"analysis","control":"run","stage_total":analyzable,"stage_processed":0,"current_store":"准备开始完整分析"});job.config=config;job.status="正在完整分析";job.processed_stores=0;job.success_stores=0
        db.commit();payload=serialize_job(job)
    _schedule_job_runner(job_id,"analysis")
    return ok(payload,"已开始连续分析全部门店")


@app.post("/api/analysis-jobs/{job_id}/retry")
def retry(job_id:int):
    with SessionLocal() as db:
        job=db.get(AnalysisJob,job_id)
        if not job: raise HTTPException(404,"任务不存在")
        links=db.scalars(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id==job_id)).all()
        retried=0
        for link in links:
            if link.status == "匹配失败":
                link.status="等待匹配";link.error_code=None;link.error_message=None;retried+=1
            elif link.status == "分析失败":
                link.status="已确认";link.error_code=None;link.error_message=None;retried+=1
        job.status="等待继续匹配" if any(x.status=="等待匹配" for x in links) else "匹配完成"
        job.failed_stores=max(0,job.failed_stores-retried);db.commit()
        return ok({"job":serialize_job(job),"retried":retried})


@app.post("/api/analysis-jobs/{job_id}/cancel")
def cancel(job_id:int):
    return set_job_status(job_id,"已取消")


@app.get("/api/analysis-jobs/{job_id}/pending-matches")
def pending(job_id:int):
    with SessionLocal() as db:
        links=db.scalars(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id==job_id,AnalysisJobStore.status=="待人工确认")).all()
        result=[]
        for link in links:
            store=db.get(Store,link.store_id)
            candidates=db.scalars(select(StoreMatchCandidate).where(StoreMatchCandidate.store_id==link.store_id).order_by(StoreMatchCandidate.score.desc())).all()
            result.append({"store":serialize_store(store),"candidates":[{"id":x.id,"amap_poi_id":x.amap_poi_id,"name":x.name,"address":x.address,"location":[x.longitude,x.latitude],"score":x.score,"reasons":x.reasons} for x in candidates]})
        return ok(result)


def _apply_match(store: Store, candidate: dict[str, Any], method: str) -> None:
    location=candidate.get("location") or [None,None]
    store.standard_name=candidate.get("name") or store.input_name
    store.amap_poi_id=candidate.get("id") or candidate.get("amap_poi_id")
    store.longitude,store.latitude=location[0],location[1]
    store.province=candidate.get("province") or store.province
    store.city=candidate.get("city") or store.city
    store.district=candidate.get("district") or store.district
    store.adcode=candidate.get("adcode") or store.adcode
    store.address=candidate.get("address") or store.address
    store.standardized_address=candidate.get("formatted_address") or candidate.get("address") or store.standardized_address
    store.geocode_level=candidate.get("level") or store.geocode_level
    store.poi_type=candidate.get("type") or store.poi_type
    store.poi_typecode=candidate.get("typecode") or store.poi_typecode
    store.match_score=int(candidate.get("score") or 0)
    store.match_status="已确认";store.location_source=method
    store.confirmation_method=method;store.confirmed_by="本地用户";store.confirmed_at=datetime.utcnow();store.last_verified_at=datetime.utcnow()


def _refresh_match_counts(db, job: AnalysisJob) -> dict[str, int]:
    links=db.scalars(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id==job.id)).all()
    statuses=[x.status for x in links]
    matched=sum(x in {"已确认","分析完成"} for x in statuses)
    pending=statuses.count("待人工确认")
    failed=sum(x in {"匹配失败","分析失败"} for x in statuses)
    remaining=sum(x in {"等待匹配","等待处理"} for x in statuses)
    job.matched_stores=matched;job.pending_stores=pending;job.failed_stores=failed
    job.processed_stores=len(statuses)-remaining
    if remaining: job.status="等待继续匹配"
    elif pending: job.status="等待人工确认"
    elif failed: job.status="匹配部分失败"
    else: job.status="匹配完成"
    return {"matched":matched,"pending":pending,"failed":failed,"remaining":remaining}


@app.post("/api/analysis-jobs/{job_id}/match-next")
async def match_next(job_id:int, body:dict[str,Any]):
    batch_size=max(1,min(50,int(body.get("batch_size") or 10)))
    auto_select_best=bool(body.get("auto_select_best",False))
    with SessionLocal() as db:
        job=db.get(AnalysisJob,job_id)
        if not job: raise HTTPException(404,"任务不存在")
        if job.status == "已取消": raise HTTPException(409,"任务已取消，请先恢复任务")
        links=db.scalars(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id==job_id,AnalysisJobStore.status.in_(["等待匹配","等待处理"])).order_by(AnalysisJobStore.id).limit(batch_size)).all()
        snapshots=[]
        for link in links:
            store=db.get(Store,link.store_id)
            snapshots.append({"link_id":link.id,"store_id":store.id,"name":store.input_name,"province":store.province or "","city":store.city or "","district":store.district or "","address":store.raw_address or store.address or "","address_only":bool(store.raw_address and store.input_name==store.raw_address)})
        job.status="正在匹配门店";db.commit()
    client=AmapClient()
    for item in snapshots:
        try:
            query={key:item[key] for key in ("name","province","city","district","address")}
            if item["address_only"]:
                raw=await client.geocode(query)
                scored=[]
                for index,candidate in enumerate(raw[:5]):
                    precise=candidate.get("level") not in {"省","市","区县","道路","未知"}
                    scored.append({**candidate,"id":f"GEO-{candidate.get('adcode') or index}-{','.join(map(str,candidate.get('location') or []))}","name":item["name"],"address":candidate.get("formatted_address") or item["address"],"score":85 if precise else 60,"auto_confirm":precise,"reasons":[f"地址解析级别：{candidate.get('level') or '未知'}"]})
            else:
                raw=await client.search_store(query)
                initial=[(candidate,evaluate_candidate(query,candidate)) for candidate in raw[:10]]
                unique_high=len([result for _,result in initial if result["score"]>=80 and not result["conflicts"]])==1
                scored=[{**candidate,**evaluate_candidate(query,candidate,unique_high_match=unique_high)} for candidate,_ in initial]
                scored.sort(key=lambda x:x["score"],reverse=True)
                for index,candidate in enumerate(scored):
                    margin=candidate["score"]-(scored[index+1]["score"] if index+1<len(scored) else 0)
                    candidate["auto_confirm"]=bool(candidate.get("auto_confirm") and margin>=15)
                    candidate["reasons"]=candidate.get("reasons") or []
            with SessionLocal() as db:
                link=db.get(AnalysisJobStore,item["link_id"]);store=db.get(Store,item["store_id"])
                for old in db.scalars(select(StoreMatchCandidate).where(StoreMatchCandidate.store_id==store.id)).all(): db.delete(old)
                for candidate in scored[:5]:
                    location=candidate.get("location") or [None,None]
                    if len(location)<2 or location[0] is None or location[1] is None: continue
                    db.add(StoreMatchCandidate(store_id=store.id,amap_poi_id=str(candidate.get("id") or ""),name=str(candidate.get("name") or store.input_name),address=str(candidate.get("address") or ""),longitude=float(location[0]),latitude=float(location[1]),score=int(candidate.get("score") or 0),reasons=list(candidate.get("reasons") or [])))
                if scored and (scored[0].get("auto_confirm") or auto_select_best):
                    method="批量自动选择最佳候选" if auto_select_best and not scored[0].get("auto_confirm") else "批量高置信度自动确认"
                    _apply_match(store,scored[0],method);link.status="已确认"
                elif scored:
                    store.match_status="待人工确认";link.status="待人工确认"
                else:
                    store.match_status="匹配失败";link.status="匹配失败";link.error_code="NO_CANDIDATE";link.error_message="高德未返回候选位置"
                db.commit()
        except Exception as exc:
            with SessionLocal() as db:
                link=db.get(AnalysisJobStore,item["link_id"]);store=db.get(Store,item["store_id"])
                link.status="匹配失败";link.error_code=getattr(exc,"code","MATCH_ERROR");link.error_message=str(exc);store.match_status="匹配失败";db.commit()
    with SessionLocal() as db:
        job=db.get(AnalysisJob,job_id);counts=_refresh_match_counts(db,job);db.commit()
        return ok({"job":serialize_job(job),**counts})


@app.post("/api/analysis-jobs/{job_id}/confirm-matches")
def confirm_matches(job_id:int, body:dict[str,Any]):
    selections=body.get("selections") or []
    if not selections: raise HTTPException(400,"未提供可确认的候选项")
    confirmed=0
    with SessionLocal() as db:
        job=db.get(AnalysisJob,job_id)
        if not job: raise HTTPException(404,"任务不存在")
        for selection in selections:
            store_id=int(selection.get("store_id") or 0);candidate_id=int(selection.get("candidate_id") or 0)
            link=db.scalar(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id==job_id,AnalysisJobStore.store_id==store_id))
            candidate=db.get(StoreMatchCandidate,candidate_id);store=db.get(Store,store_id)
            if not link or not candidate or candidate.store_id!=store_id or not store: continue
            _apply_match(store,{"id":candidate.amap_poi_id,"name":candidate.name,"address":candidate.address,"location":[candidate.longitude,candidate.latitude],"score":candidate.score},"批量人工选择候选")
            link.status="已确认";link.error_code=None;link.error_message=None;confirmed+=1
        counts=_refresh_match_counts(db,job);db.commit()
        return ok({"confirmed":confirmed,"job":serialize_job(job),**counts},"候选门店已确认")


def _schedule_job_runner(job_id:int,stage:str) -> None:
    active=JOB_RUNNERS.get(job_id)
    if active and not active.done():
        return
    task=asyncio.create_task(_run_match_all_background(job_id) if stage=="match" else _run_analysis_all_background(job_id))
    JOB_RUNNERS[job_id]=task
    task.add_done_callback(lambda finished,task_job_id=job_id: JOB_RUNNERS.pop(task_job_id,None) if JOB_RUNNERS.get(task_job_id) is finished else None)


def _runtime_control(job_id:int) -> str:
    with SessionLocal() as db:
        job=db.get(AnalysisJob,job_id)
        return str((job.config or {}).get("control") or "run") if job else "stop"


def _update_runtime(job_id:int,**updates:Any) -> None:
    with SessionLocal() as db:
        job=db.get(AnalysisJob,job_id)
        if not job:return
        config=dict(job.config or {});config.update(updates);job.config=config;db.commit()


async def _run_match_all_background(job_id:int) -> None:
    try:
        with SessionLocal() as db:
            job=db.get(AnalysisJob,job_id)
            if not job:return
            pending=db.scalars(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id==job_id,AnalysisJobStore.status=="待人工确认")).all()
            for link in pending:
                candidate=db.scalar(select(StoreMatchCandidate).where(StoreMatchCandidate.store_id==link.store_id).order_by(StoreMatchCandidate.score.desc()).limit(1))
                store=db.get(Store,link.store_id)
                if candidate and store:
                    _apply_match(store,{"id":candidate.amap_poi_id,"name":candidate.name,"address":candidate.address,"location":[candidate.longitude,candidate.latitude],"score":candidate.score},"批量自动选择最佳候选")
                    link.status="已确认"
                else:
                    link.status="等待匹配"
            _refresh_match_counts(db,job);db.commit()
        while True:
            control=_runtime_control(job_id)
            if control in {"pause","stop"}:
                with SessionLocal() as db:
                    job=db.get(AnalysisJob,job_id)
                    if job:job.status="已暂停" if control=="pause" else "已结束";db.commit()
                return
            with SessionLocal() as db:
                job=db.get(AnalysisJob,job_id)
                if not job:return
                link=db.scalar(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id==job_id,AnalysisJobStore.status.in_(["等待匹配","等待处理"])).order_by(AnalysisJobStore.id).limit(1))
                if not link:
                    counts=_refresh_match_counts(db,job)
                    config=dict(job.config or {});config.update({"control":"idle","stage_processed":job.total_stores,"current_store":""});job.config=config
                    job.status="匹配部分失败" if counts["failed"] else "匹配完成";db.commit();return
                store=db.get(Store,link.store_id)
                config=dict(job.config or {});config.update({"active_stage":"match","control":"run","stage_total":job.total_stores,"current_store":store.input_name if store else ""});job.config=config;job.status="正在匹配门店";db.commit()
            result=await match_next(job_id,{"batch_size":1,"auto_select_best":True})
            matched_data=result.get("data") or {}
            matched_job=matched_data.get("job") or {}
            _update_runtime(job_id,stage_processed=int(matched_job.get("processed_stores") or 0))
            await asyncio.sleep(0)
    except Exception as exc:
        with SessionLocal() as db:
            job=db.get(AnalysisJob,job_id)
            if job:
                config=dict(job.config or {});config.update({"control":"idle","current_store":""});job.config=config;job.status="任务异常";db.commit()
        return


async def _analyze_background_store(job_id:int,store_id:int,categories:list[str],radii:list[int]) -> dict[str,Any]:
    poi_status=await _run_batch_poi_analysis(job_id,store_id,categories,radii)
    await _run_business_analysis(store_id,radii,analysis_job_id=job_id)
    return poi_status


async def _run_analysis_all_background(job_id:int) -> None:
    any_truncated=False
    try:
        while True:
            control=_runtime_control(job_id)
            if control in {"pause","stop"}:
                with SessionLocal() as db:
                    job=db.get(AnalysisJob,job_id)
                    if job:job.status="已暂停" if control=="pause" else "已结束";db.commit()
                return
            with SessionLocal() as db:
                job=db.get(AnalysisJob,job_id)
                if not job:return
                categories=[value for value in (job.config or {}).get("categories",[]) if value in CATEGORIES] or ["住宅小区","幼儿园","小学"]
                radii=[int(value) for value in ((job.config or {}).get("radii") or [500])]
                links=db.scalars(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id==job_id,AnalysisJobStore.status=="已确认").order_by(AnalysisJobStore.id).limit(MAX_CONCURRENT_STORES)).all()
                if not links:
                    all_links=db.scalars(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id==job_id)).all()
                    success=sum(link.status=="分析完成" for link in all_links);failed=sum(link.status in {"匹配失败","分析失败"} for link in all_links)
                    config=dict(job.config or {});config.update({"control":"idle","stage_processed":success+sum(link.status=="分析失败" for link in all_links),"current_store":""});job.config=config
                    job.processed_stores=success+sum(link.status=="分析失败" for link in all_links);job.success_stores=success;job.failed_stores=failed;job.truncated=any_truncated
                    job.status="部分完成" if failed else "已完成";db.commit();return
                store_ids=[link.store_id for link in links]
                names=[db.get(Store,store_id).input_name for store_id in store_ids if db.get(Store,store_id)]
                config=dict(job.config or {});config.update({"active_stage":"analysis","control":"run","current_store":"、".join(names)});job.config=config;job.status="正在完整分析";db.commit()
            outcomes=await asyncio.gather(*[_analyze_background_store(job_id,store_id,categories,radii) for store_id in store_ids],return_exceptions=True)
            with SessionLocal() as db:
                job=db.get(AnalysisJob,job_id)
                for store_id,outcome in zip(store_ids,outcomes):
                    if isinstance(outcome,Exception):
                        link=db.scalar(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id==job_id,AnalysisJobStore.store_id==store_id))
                        if link:link.status="分析失败";link.error_code=getattr(outcome,"code","ANALYSIS_ERROR");link.error_message=str(outcome)
                    else:
                        any_truncated=any_truncated or bool(outcome.get("truncated"))
                all_links=db.scalars(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id==job_id)).all()
                processed=sum(link.status in {"分析完成","分析失败"} for link in all_links);success=sum(link.status=="分析完成" for link in all_links);failed=sum(link.status in {"匹配失败","分析失败"} for link in all_links)
                config=dict(job.config or {});config["stage_processed"]=processed;job.config=config;job.processed_stores=processed;job.success_stores=success;job.failed_stores=failed;db.commit()
            await asyncio.sleep(0)
    except Exception:
        with SessionLocal() as db:
            job=db.get(AnalysisJob,job_id)
            if job:
                config=dict(job.config or {});config.update({"control":"idle","current_store":""});job.config=config;job.status="任务异常";db.commit()
        return


@app.post("/api/analysis-jobs/{job_id}/business-district-analysis")
async def create_batch_business_analysis(job_id: int, body: BusinessAnalysisBody):
    with SessionLocal() as db:
        job = db.get(AnalysisJob, job_id)
        if not job:
            raise HTTPException(404, "任务不存在")
        links = db.scalars(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id == job_id)).all()
        stores = {link.store_id: db.get(Store, link.store_id) for link in links}
        store_ids = [link.store_id for link in links if stores[link.store_id] and stores[link.store_id].longitude is not None and link.status in {"已确认","已完成","分析完成","分析失败"}]
        pending_count = len(links) - len(store_ids)
        if not store_ids:
            raise HTTPException(400, "当前任务没有已确认坐标的门店，请先完成批量匹配")
        categories = [value for value in (job.config or {}).get("categories", []) if value in CATEGORIES]
        if not categories:
            categories = ["住宅小区", "幼儿园", "小学"]
        job.status = "正在生成商圈画像"; job.processed_stores = 0; db.commit()
    completed, failed, warnings = 0, [], []
    any_truncated = False
    for store_id in store_ids:
        try:
            poi_status = await _run_batch_poi_analysis(job_id, store_id, categories, body.radii)
            any_truncated = any_truncated or poi_status["truncated"]
            warnings.extend({"store_id": store_id, "message": message} for message in poi_status["warnings"])
            await _run_business_analysis(store_id, body.radii, analysis_job_id=job_id)
            completed += 1
        except Exception as exc:
            failed.append({"store_id": store_id, "message": str(exc)})
    with SessionLocal() as db:
        job = db.get(AnalysisJob, job_id)
        job.processed_stores = completed + len(failed)
        job.success_stores = completed
        job.failed_stores = len(failed)
        job.pending_stores = pending_count
        job.truncated = any_truncated
        job.status = "已完成" if not failed and not pending_count else "部分完成"
        db.commit()
    return ok({"job_id": job_id, "completed": completed, "failed": failed, "warnings": warnings, "pending": pending_count,"status": "已完成" if not failed and not pending_count else "部分完成"})


async def _run_batch_poi_analysis(job_id: int, store_id: int, categories: list[str], radii: list[int]) -> dict[str, Any]:
    """Collect the task-configured POIs for one imported store without mixing historical jobs."""
    max_radius = max(radii)
    with SessionLocal() as db:
        store = db.get(Store, store_id)
        if not store or store.longitude is None or store.latitude is None:
            raise HTTPException(400, "请先确认门店位置")
        existing = db.scalars(select(PoiResult).where(
            PoiResult.analysis_job_id == job_id,
            PoiResult.store_id == store_id,
            PoiResult.poi_category.in_(categories),
        )).all()
        completed_categories = {item.poi_category for item in existing}
        longitude, latitude = store.longitude, store.latitude

    client = AmapClient()
    collected: list[dict[str, Any]] = []
    warnings: list[str] = []
    truncated = False
    for category in categories:
        if category in completed_categories:
            continue
        try:
            items, cut = await client.around((longitude, latitude), category, max_radius)
            truncated = truncated or cut
            collected.extend({**item, "category": category} for item in items if item["distance"] <= max_radius)
        except AmapServiceError as exc:
            warnings.append(f"{category}查询失败：{exc.message}")

    with SessionLocal() as db:
        for item in deduplicate_pois(collected):
            exists = db.scalar(select(PoiResult.id).where(
                PoiResult.analysis_job_id == job_id,
                PoiResult.store_id == store_id,
                PoiResult.amap_poi_id == item["id"],
                PoiResult.poi_category == item["category"],
            ).limit(1))
            if exists:
                continue
            db.add(PoiResult(
                analysis_job_id=job_id, store_id=store_id, amap_poi_id=item["id"], parent_poi_id=item.get("parent"),
                name=item["name"], poi_category=item["category"], poi_type=item.get("type"), poi_typecode=item.get("typecode"),
                longitude=item["location"][0], latitude=item["location"][1], straight_distance_m=item["distance"],
                province=item.get("province"), city=item.get("city"), district=item.get("district"), adcode=item.get("adcode"),
                address=item.get("address"), business_area=item.get("business_area"), distance_bucket=distance_bucket(item["distance"], radii),
                search_keyword=item["category"] if item["category"] == "竞品门店" else None,
                search_typecodes=CATEGORIES.get(item["category"], ("", ""))[0], search_radius=max_radius,
            ))
        db.commit()
    return {"truncated": truncated, "warnings": warnings}


def _job_poi_rows(db, job: AnalysisJob, store_id: int) -> list[PoiResult]:
    categories = [value for value in (job.config or {}).get("categories", []) if value in CATEGORIES]
    query = select(PoiResult).where(
        PoiResult.analysis_job_id == job.id,
        PoiResult.store_id == store_id,
        PoiResult.excluded == False,
    )
    if categories:
        query = query.where(PoiResult.poi_category.in_(categories))
    return list(db.scalars(query.order_by(PoiResult.straight_distance_m)).all())


def _summarize_pois(items: list[PoiResult]) -> dict[str, Any]:
    by_category: dict[str, int] = {}
    by_distance: dict[str, int] = {}
    for item in items:
        by_category[item.poi_category] = by_category.get(item.poi_category, 0) + 1
        by_distance[item.distance_bucket] = by_distance.get(item.distance_bucket, 0) + 1
    return {"total": len(items), "by_category": by_category, "by_distance": by_distance}


@app.get("/api/analysis-jobs/{job_id}/stores")
def get_batch_job_stores(job_id: int):
    with SessionLocal() as db:
        job = db.get(AnalysisJob, job_id)
        if not job:
            raise HTTPException(404, "任务不存在")
        links = db.scalars(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id == job_id).order_by(AnalysisJobStore.id)).all()
        result = []
        for link in links:
            store = db.get(Store, link.store_id)
            if not store:
                continue
            items = _job_poi_rows(db, job, store.id)
            analysis = db.scalar(select(BusinessDistrictAnalysis).where(
                BusinessDistrictAnalysis.analysis_job_id == job_id,
                BusinessDistrictAnalysis.store_id == store.id,
            ).order_by(BusinessDistrictAnalysis.created_at.desc()).limit(1))
            result.append({"store": serialize_store(store), "status": link.status, "poi_summary": _summarize_pois(items), "has_profile": analysis is not None, "error_message": link.error_message})
        return ok(result)


@app.get("/api/analysis-jobs/{job_id}/stores/{store_id}")
def get_batch_store_detail(job_id: int, store_id: int):
    with SessionLocal() as db:
        job = db.get(AnalysisJob, job_id)
        link = db.scalar(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id == job_id, AnalysisJobStore.store_id == store_id))
        store = db.get(Store, store_id)
        if not job or not link or not store:
            raise HTTPException(404, "任务中的门店不存在")
        items = _job_poi_rows(db, job, store_id)
        analysis = db.scalar(select(BusinessDistrictAnalysis).where(
            BusinessDistrictAnalysis.analysis_job_id == job_id,
            BusinessDistrictAnalysis.store_id == store_id,
        ).order_by(BusinessDistrictAnalysis.created_at.desc()).limit(1))
        return ok({
            "job": serialize_job(job), "store": serialize_store(store), "status": link.status,
            "pois": [serialize_poi(item) for item in items], "poi_summary": _summarize_pois(items),
            "analysis": _serialize_business_analysis(analysis) if analysis else None,
            "disclaimer": "POI 与画像均限定在本批量任务内；潜在人群和消费环境属于周边设施代理推断，不代表真实人口或销售数据。",
        })


@app.get("/api/analysis-jobs/{job_id}/business-district-results")
def get_batch_business_results(job_id: int):
    with SessionLocal() as db:
        job = db.get(AnalysisJob, job_id)
        if not job:
            raise HTTPException(404, "任务不存在")
        items = db.scalars(select(BusinessDistrictAnalysis).where(BusinessDistrictAnalysis.analysis_job_id == job_id).order_by(BusinessDistrictAnalysis.created_at.desc())).all()
        result = []
        seen: set[int] = set()
        for item in items:
            if item.store_id in seen:
                continue
            seen.add(item.store_id)
            store = db.get(Store, item.store_id)
            pois = _job_poi_rows(db, job, item.store_id)
            result.append({**_serialize_business_analysis(item), "store": serialize_store(store) if store else None, "poi_summary": _summarize_pois(pois)})
        result.sort(key=lambda value: value["fit"]["score"], reverse=True)
        return ok(result)


@app.get("/api/analysis-jobs/{job_id}/stores/{store_id}/export")
def export_batch_store(job_id: int, store_id: int):
    with SessionLocal() as db:
        job = db.get(AnalysisJob, job_id)
        link = db.scalar(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id == job_id, AnalysisJobStore.store_id == store_id))
        store = db.get(Store, store_id)
        if not job or not link or not store:
            raise HTTPException(404, "任务中的门店不存在")
        pois = _job_poi_rows(db, job, store_id)
        analysis = db.scalar(select(BusinessDistrictAnalysis).where(
            BusinessDistrictAnalysis.analysis_job_id == job_id,
            BusinessDistrictAnalysis.store_id == store_id,
        ).order_by(BusinessDistrictAnalysis.created_at.desc()).limit(1))
    wb = Workbook(); overview = wb.active; overview.title = "门店概览"
    overview.append(["项目", "内容"])
    for label, value in [
        ("任务编号", job_id), ("门店编号", store.user_code), ("输入门店名称", store.input_name), ("高德标准门店名称", store.standard_name),
        ("匹配状态", store.match_status), ("匹配分", store.match_score), ("城市", store.city), ("区县", store.district),
        ("地址", store.address), ("经度", store.longitude), ("纬度", store.latitude), ("任务状态", link.status),
        ("POI总数", len(pois)), ("搜索半径", str((job.config or {}).get("radii", []))), ("POI分类", "、".join((job.config or {}).get("categories", []))),
        ("数据说明", "潜在人群和消费环境属于周边设施代理推断，不代表真实人口、客流或销售数据。"),
    ]:
        overview.append([label, safe_excel(value)])
    style_sheet(overview)
    details = wb.create_sheet("POI明细")
    details.append(["POI分类", "POI名称", "类型", "地址", "直线距离", "距离分层", "经度", "纬度", "高德POI ID"])
    for item in pois:
        details.append([safe_excel(value) for value in [item.poi_category, item.name, item.poi_type, item.address, item.straight_distance_m, item.distance_bucket, item.longitude, item.latitude, item.amap_poi_id]])
    style_sheet(details)
    if analysis:
        append_business_sheets(wb, [analysis], {store.id: store})
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=batch_store_{job_id}_{store_id}.xlsx"})


@app.get("/api/analysis-jobs/{job_id}/export")
def export_job(job_id: int):
    with SessionLocal() as db:
        job=db.get(AnalysisJob,job_id)
        if not job: raise HTTPException(404,"任务不存在")
        links=db.scalars(select(AnalysisJobStore).where(AnalysisJobStore.analysis_job_id==job_id)).all()
        stores=[db.get(Store,x.store_id) for x in links]
        pois=db.scalars(select(PoiResult).where(PoiResult.analysis_job_id==job_id)).all()
        analyses=db.scalars(select(BusinessDistrictAnalysis).where(BusinessDistrictAnalysis.analysis_job_id==job_id)).all()
        candidate_counts={link.store_id:len(db.scalars(select(StoreMatchCandidate).where(StoreMatchCandidate.store_id==link.store_id)).all()) for link in links}
    wb=Workbook();ws=wb.active;ws.title="门店汇总"
    ws.append(["用户门店编号","用户输入门店名称","高德标准门店名称","匹配状态","匹配分数","省份","城市","区县","地址","经度","纬度","高德POI ID","是否存在截断","最后分析时间"])
    for s in stores: ws.append([safe_excel(x) for x in [s.user_code,s.input_name,s.standard_name,s.match_status,s.match_score,s.province,s.city,s.district,s.address,s.longitude,s.latitude,s.amap_poi_id,"是" if job.truncated else "否",job.updated_at]])
    style_sheet(ws)
    ws2=wb.create_sheet("POI明细");ws2.append(["用户输入门店名称","高德标准门店名称","门店POI ID","门店经度","门店纬度","POI分类","POI名称","POI type","POI typecode","POI ID","省份","城市","区县","地址","POI经度","POI纬度","直线距离","步行距离","步行时间","距离分层","搜索关键词","搜索typecodes","搜索半径","查询时间","是否人工修正","是否排除"])
    store_by_id={s.id:s for s in stores}
    for p in pois:
        s=store_by_id.get(p.store_id);ws2.append([safe_excel(x) for x in [s.input_name,s.standard_name,s.amap_poi_id,s.longitude,s.latitude,p.poi_category,p.name,p.poi_type,p.poi_typecode,p.amap_poi_id,p.province,p.city,p.district,p.address,p.longitude,p.latitude,p.straight_distance_m,p.walking_distance_m,p.walking_duration_s,p.distance_bucket,p.search_keyword,p.search_typecodes,p.search_radius,p.created_at,p.manually_corrected,p.excluded]])
    style_sheet(ws2)
    link_by_store={link.store_id:link for link in links}
    for title,headers in [("待确认门店",["原始门店名称","城市","候选数量","匹配状态","原因"]),("失败记录",["门店","处理阶段","错误代码","中文错误说明","是否可重试","最后重试时间"]),("搜索配置",["搜索半径","POI分类","typecodes","关键词","查询时间","坐标体系","数据来源说明"])]:
        sheet=wb.create_sheet(title);sheet.append(headers)
        if title=="待确认门店":
            for store in stores:
                link=link_by_store.get(store.id)
                if link and link.status=="待人工确认": sheet.append([safe_excel(store.input_name),safe_excel(store.city),candidate_counts.get(store.id,0),store.match_status,"请人工选择正确候选"])
        elif title=="失败记录":
            for store in stores:
                link=link_by_store.get(store.id)
                if link and link.status in {"匹配失败","分析失败"}: sheet.append([safe_excel(store.input_name),"门店匹配" if link.status=="匹配失败" else "画像分析",link.error_code,safe_excel(link.error_message),"是",link.updated_at])
        elif title=="搜索配置": sheet.append([str(job.config.get("radii",[])),",".join(job.config.get("categories",[])),"见分类配置","竞品按关键词",job.updated_at,"GCJ-02","演示模式" if os.getenv("ENABLE_MOCK_AMAP","true").lower()=="true" else "高德 Web 服务 API"])
        style_sheet(sheet)
    append_business_sheets(wb, analyses, {s.id: s for s in stores})
    buf=io.BytesIO();wb.save(buf);buf.seek(0)
    return StreamingResponse(buf,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f"attachment; filename=poi_analysis_{job_id}.xlsx"})


@app.get("/api/analysis-jobs/{job_id}/business-district-export")
def export_business_results(job_id: int):
    with SessionLocal() as db:
        analyses = db.scalars(select(BusinessDistrictAnalysis).where(BusinessDistrictAnalysis.analysis_job_id == job_id)).all()
        if not analyses:
            raise HTTPException(404, "该任务暂无商圈分析结果")
        store_map = {item.store_id: db.get(Store, item.store_id) for item in analyses}
    wb = Workbook(); wb.remove(wb.active)
    append_business_sheets(wb, analyses, store_map)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=business_district_{job_id}.xlsx"})


def append_business_sheets(wb: Workbook, analyses: list[BusinessDistrictAnalysis], store_map: dict[int, Store | None]) -> None:
    summary = wb.create_sheet("商圈汇总")
    summary.append(["门店编号","输入门店名称","高德标准门店名称","城市","区县","地址","经度","纬度","定位来源","匹配分","商圈名称","商圈名称来源","商圈类型","商圈类型可信度","商圈能级","商圈能级得分","评级方式","业务适配度","适配等级","竞争压力","商圈可信度","主要优势","主要不足","数据警告","分析时间","算法版本"])
    for item in analyses:
        store = store_map.get(item.store_id)
        values = [
            store.user_code if store else None, store.input_name if store else None, store.standard_name if store else None,
            store.city if store else None, store.district if store else None, store.address if store else None,
            store.longitude if store else None, store.latitude if store else None, item.location_source,
            store.match_score if store else None, item.business_area_name, item.business_area_source,
            item.business_district_type, item.type_confidence, item.level, item.level_score, item.level_mode,
            item.fit_score, item.fit_level, item.competition_level, item.confidence_level,
            "；".join(item.strengths or []), "；".join(item.weaknesses or []), "；".join(item.warning_messages or []),
            item.created_at, item.analysis_version,
        ]
        summary.append([safe_excel(value) for value in values])
    style_sheet(summary)

    details = wb.create_sheet("商圈特征明细")
    details.append(["门店","半径","POI分类","POI数量","POI密度","最近距离","是否截断","数据是否完整","特征标准化得分","特征权重","特征贡献值"])
    for item in analyses:
        store = store_map.get(item.store_id)
        for radius, layer in (item.feature_vector or {}).get("layers", {}).items():
            counts = layer.get("counts", {})
            for category, count in counts.items():
                details.append([safe_excel(store.input_name if store else item.store_id), int(radius), category, count, layer.get("density"), layer.get("nearest", {}).get(category), "是" if (item.truncation_flags or {}).get("any") else "否", "否" if item.warning_messages else "是", None, None, None])
    style_sheet(details)

    audience = wb.create_sheet("潜在人群画像")
    audience.append(["门店","主要潜在人群","年龄段倾向","消费环境判断","消费环境指数","商场档次线索","商场样本数","画像可信度","推断依据","重要限制"])
    for item in analyses:
        store = store_map.get(item.store_id)
        profile = (item.feature_vector or {}).get("audience_profile") or {}
        groups = profile.get("primary_groups") or []
        consumption = profile.get("consumption_power") or {}
        mall = profile.get("mall_profile") or {}
        audience.append([safe_excel(value) for value in [
            store.input_name if store else item.store_id,
            "；".join(str(group.get("label") or "") for group in groups),
            "；".join(str(group.get("age_range") or "") for group in groups),
            consumption.get("level"), consumption.get("index"), mall.get("level"), mall.get("sample_count"),
            profile.get("confidence"), "；".join(profile.get("evidence") or []), "；".join(profile.get("limitations") or []),
        ]])
    style_sheet(audience)

    config = wb.create_sheet("商圈评分配置")
    config.append(["配置项","配置内容"])
    cfg = public_config()
    for key in ("type_rules", "level_weights", "fit_weights", "default_radii", "analysis_version", "poi_config_version", "weight_version"):
        config.append([key, safe_excel(str(cfg.get(key)))])
    config.append(["坐标体系", "GCJ-02"])
    config.append(["数据来源说明", DISCLAIMER])
    style_sheet(config)


def set_job_status(job_id:int,status:str):
    with SessionLocal() as db:
        job=db.get(AnalysisJob,job_id)
        if not job: raise HTTPException(404,"任务不存在")
        job.status=status;db.commit();return ok(serialize_job(job))


def serialize_store(s:Store):
    return {"id":s.id,"input_name":s.input_name,"standard_name":s.standard_name,"amap_poi_id":s.amap_poi_id,"longitude":s.longitude,"latitude":s.latitude,"province":s.province,"city":s.city,"district":s.district,"adcode":s.adcode,"address":s.address,"raw_address":s.raw_address,"standardized_address":s.standardized_address,"geocode_level":s.geocode_level,"location_source":s.location_source,"match_score":s.match_score,"match_status":s.match_status,"confirmation_method":s.confirmation_method}


def serialize_poi(p:PoiResult):
    return {"id":p.amap_poi_id,"name":p.name,"category":p.poi_category,"type":p.poi_type,"typecode":p.poi_typecode,"address":p.address,"business_area":p.business_area,"distance":p.straight_distance_m,"walking_distance":p.walking_distance_m,"location":[p.longitude,p.latitude],"distance_bucket":p.distance_bucket}


def serialize_job(j:AnalysisJob):
    config=j.config or {}
    stage=str(config.get("active_stage") or ("analysis" if j.success_stores else "match"))
    stage_total=int(config.get("stage_total") or (j.matched_stores if stage=="analysis" and j.matched_stores else j.total_stores) or 0)
    if "stage_processed" in config:
        stage_processed=int(config.get("stage_processed") or 0)
    elif stage=="analysis":
        stage_processed=int(j.processed_stores or (j.success_stores+j.failed_stores))
    else:
        stage_processed=int(j.processed_stores or 0)
    return {"id":j.id,"filename":j.filename,"status":j.status,"total_stores":j.total_stores,"processed_stores":j.processed_stores,"matched_stores":j.matched_stores,"pending_stores":j.pending_stores,"success_stores":j.success_stores,"failed_stores":j.failed_stores,"truncated":j.truncated,"config":config,"stage":stage,"control":config.get("control") or "idle","stage_total":stage_total,"stage_processed":stage_processed,"progress_percent":round(stage_processed/max(1,stage_total)*100),"current_store":config.get("current_store") or "","created_at":j.created_at.isoformat(),"updated_at":j.updated_at.isoformat()}


def style_sheet(ws):
    ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions
    for cell in ws[1]: cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="08745B")
    for col in range(1,ws.max_column+1):
        values=[len(str(ws.cell(r,col).value or "")) for r in range(1,min(ws.max_row,100)+1)]
        ws.column_dimensions[get_column_letter(col)].width=min(max(max(values,default=8)+2,10),36)


@app.exception_handler(HTTPException)
async def http_error(request:Request,exc:HTTPException):
    from fastapi.responses import JSONResponse
    detail=exc.detail
    if isinstance(detail,dict): code=detail.get("error_code","REQUEST_ERROR");message=detail.get("message","请求失败");details=detail
    else: code="REQUEST_ERROR";message=str(detail);details={}
    return JSONResponse(status_code=exc.status_code,content={"success":False,"error_code":code,"message":message,"details":details,"request_id":request.state.request_id})


@app.exception_handler(Exception)
async def unhandled(request:Request,exc:Exception):
    from fastapi.responses import JSONResponse
    return JSONResponse(status_code=500,content={"success":False,"error_code":"INTERNAL_ERROR","message":"服务处理失败，请使用请求编号联系管理员","details":{},"request_id":request.state.request_id})
